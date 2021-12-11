from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
from bokeh.plotting import figure
from bokeh.palettes import all_palettes

def import_data(file_n, colx, coly, cell_n):
    wb = load_workbook(os.getcwd()+file_n)
    ws = wb.active

    range_x = ws[colx+'2':colx+'2049']
    x = []
    range_y = ws[coly+'2':coly+'2049']
    y = []

    for cell in range_x:
        for e in cell:
            x.append(e.value)

    if cell_n == None:
        for cell in range_y:
            for e in cell:
                y.append(e.value)
    else:
        for cell in range_y:
            for e in cell:
                y.append(e.value/ws[cell_n].value) #normalizuję dane

    data = {'x': x, 'y': y}
    return data

#pobieram dane o źródle światła białego
white_spectra = import_data('\\source\\white.xlsx', 'A', 'B', 'B645')

#pobieram dane o diodach LED
leds = {}
for led in [['red', 'A', 'B', 'B876'], ['orange', 'E', 'F', 'F746'], ['yellow', 'I', 'J', 'J687' ],
['green', 'M', 'N', 'N637'], ['blue', 'Q', 'R', 'R329'], ['uv', 'U', 'V', 'V162']]:
    leds[f'{led[0]}_spectrum'] = import_data('\\source\\leds.xlsx', led[1], led[2], led[3])

#pobieram dane o olejach 
abs = {}
trans = {}
for oil in [['diesel', '\\diesel\\diesel.xlsx'], ['castrol_edge_0w20', '\\engine\\Castrol EDGE 0W-20.xlsx'],
['castrol_edge_0w30', '\\engine\\Castrol EDGE 0W-30.xlsx'], ['castrol_magnatec', '\\engine\\Castrol MAGNATEC.xlsx'],
['comma_pro-nrg', '\\engine\\COMMA Pro-NRG.xlsx'], ['comma_xtech', '\\engine\\COMMA Xtech.xlsx'],
['elf_evolution', '\\engine\\elf EVOLUTION.xlsx'], ['millers_oils', '\\engine\\MILLERS OILS.xlsx'],
['mobil_super_2000', '\\engine\\Mobil Super 2000.xlsx'], ['motul_specific', '\\engine\\MOTUL Specific.xlsx'],
['comma_asw', '\\gear\\COMMA ASW.xlsx'], ['febi_bilstein_axle_drive', '\\gear\\febi bilstein axle drive.xlsx'],
['febi_bilstein_gear_oil', '\\gear\\febi bilstein gear oil.xlsx'],
['olive', '\\cooking\\olive.xlsx'], ['rapeseed', '\\cooking\\rapeseed.xlsx'], ['sunflower', '\\cooking\\sunflower.xlsx']]:
    abs[f'{oil[0]}_abs'] = import_data(oil[1], 'A', 'B', None)
    trans[f'{oil[0]}_trans'] = import_data(oil[1], 'E', 'F', None)

def set_fig_properties(fig):
    fig.background_fill_color = all_palettes['Greys'][9][7]
    fig.ygrid.grid_line_color = all_palettes['Greys'][9][6]
    fig.xgrid.grid_line_color = all_palettes['Greys'][9][6]
    fig.title.text_font_size = '16px'
    fig.title.text_font = 'Calibri Light'
    fig.yaxis.axis_label_text_font_size = '14px'
    fig.xaxis.axis_label_text_font_size = '14px'
    fig.yaxis.axis_label_text_font = 'Calibri Light'
    fig.xaxis.axis_label_text_font = 'Calibri Light'

#tworzę wykres źródła światła białego
fig_white = figure(x_range = (300,900), x_axis_label = 'Wavelength [nm]', y_axis_label = 'Light intensity [a.u.]',
             title = 'White light source spectrum', width = 900, height = 580)
set_fig_properties(fig_white)

white_line = fig_white.line('x', 'y', source = white_spectra, name = 'white_line', color = all_palettes['Dark2'][6][5])
white_patch = fig_white.patch('x', source = white_spectra, name = 'white_patch', color = all_palettes['Dark2'][6][5], alpha = 0.2)

#tworzę wykres do matchowania
fig_match = figure(x_range = (300,1050), y_range = (0, 4), x_axis_label = 'Wavelength [nm]', y_axis_label = 'Light intensity [a.u.]',
             title = 'Selected spectra', width = 900, height = 580)
set_fig_properties(fig_match)

#tworzę wykres diod LED
fig_LED = figure(x_range = (350,750), x_axis_label = 'Wavelength [nm]', y_axis_label = 'Light intensity [a.u.]',
             title = 'LED diodes\' spectra', width = 900, height = 580)
set_fig_properties(fig_LED)

leds_lines = {}
m_leds_lines = {}
for line in [['red', 'RED', all_palettes['Spectral'][11][10]], ['orange', 'ORANGE', all_palettes['Set2'][6][1]], ['yellow', 'YELLOW', all_palettes['Set2'][6][5]], 
['green', 'GREEN', all_palettes['PuBuGn'][5][0]], ['blue', 'BLUE', all_palettes['Spectral'][11][1]], ['uv', 'UV', all_palettes['Accent'][3][1]]]:
    leds_lines[f'{line[0]}_line'] = fig_LED.line('x', 'y', source = leds[f'{line[0]}_spectrum'], name = f'{line[0]}_line', legend_label = f'{line[1]}', color = line[2])
    leds_lines[f'{line[0]}_patch'] = fig_LED.patch('x', source = leds[f'{line[0]}_spectrum'], name = f'{line[0]}_patch', legend_label = f'{line[1]}', color = line[2], alpha = 0.2)
    m_leds_lines[f'm_{line[0]}_line'] = fig_match.line('x', 'y', source = leds[f'{line[0]}_spectrum'], name = f'{line[0]}_line', legend_label = f'{line[1]}', color = line[2])

#tworzę wykres dla widm absorpcyjnych
fig_abs = figure(x_range = (400,800), y_range = (0, 4), x_axis_label = 'Wavelength [nm]', y_axis_label = 'Absorbance [OD]',
             title = 'Oil\'s absorption spectrum', width = 900, height = 296)
set_fig_properties(fig_abs)

fig_trans = figure(x_range = (400,800), y_range = (0, 100), x_axis_label = 'Wavelength [nm]', y_axis_label = 'Transmitance [%]',
             title = 'Oil\'s transmission spectrum', width = 900, height = 296)
set_fig_properties(fig_trans)

abs_lines = {}
trans_lines = {}
m_abs_lines = {}
for line in [['diesel', 'DIESEL', all_palettes['Set2'][3][2]], ['castrol_edge_0w20', 'ENGINE_CASTROL_EDGE_0W-20', all_palettes['Set2'][3][1]],
['castrol_edge_0w30', 'ENGINE_CASTROL_EDGE_0W-30', all_palettes['Set2'][3][0]], ['castrol_magnatec', 'CASTROL_MAGNATEC', all_palettes['Set2'][4][3]], 
['comma_pro-nrg', 'COMMA PRO-NRG', all_palettes['BrBG'][11][1]], ['comma_xtech', 'COMMA XTECH', all_palettes['BrBG'][11][9]],
['elf_evolution', 'ELF EVOLUTION', all_palettes['BuPu'][8][0]], ['millers_oils', 'MILLERS OILS', all_palettes['Set3'][12][9]],
['mobil_super_2000', 'MOBIL SUPER 2000', all_palettes['Set3'][12][3]], ['motul_specific', 'MOTUL Specific', all_palettes['Bokeh'][3][1]], ['comma_asw', 'COMMA ASW', all_palettes['Bokeh'][5][4]],
['febi_bilstein_axle_drive', 'FEBI BILSTEIN AXLE DRIVE', all_palettes['Set3'][12][0]], ['febi_bilstein_gear_oil', 'FEBI BILSTEIN GEAR OIL', all_palettes['Set3'][12][4]],
['olive', 'OLIVE', all_palettes['Reds'][5][0]], ['rapeseed', 'RAPESEED', all_palettes['Bokeh'][6][5]], ['sunflower', 'SUNFLOWER', all_palettes['Cividis'][3][1]]]:
    abs_lines[f'{line[0]}_abs_line'] = fig_abs.line('x', 'y', source = abs[f'{line[0]}_abs'], name = f'{line[0]}_abs_line', legend_label = f'{line[1]}', color = line[2])
    trans_lines[f'{line[0]}_trans_line'] = fig_trans.line('x', 'y', source = trans[f'{line[0]}_trans'], name = f'{line[0]}_trans_line', legend_label = f'{line[1]}', color = line[2])
    m_abs_lines[f'm_{line[0]}_abs_line'] = fig_match.line('x', 'y', source = abs[f'{line[0]}_abs'], name = f'{line[0]}_abs_line', legend_label = f'{line[1]}', color = line[2])

#ustawienia legend
fig_LED.legend.click_policy = "hide"
fig_abs.legend.visible = False
fig_trans.legend.visible = False  
fig_match.legend.visible = False 

for line in m_leds_lines:
    m_leds_lines[line].visible = False

for line in m_abs_lines:
    m_abs_lines[line].visible = False

