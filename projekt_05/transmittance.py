import numpy as np
from sklearn.linear_model import LinearRegression
from bokeh.models import ColumnDataSource, DataTable, TableColumn, HTMLTemplateFormatter, Div
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
from collections import defaultdict
from bokeh.plotting import figure
from bokeh.palettes import all_palettes

import figs

# div_table_white = Div(text = 'Predicted transmittance values for VIS LED diodes based on white light source results:')

# def get_max_lambda(data):
#     y = np.array([])
#     x = np.array([])
#     y = np.append(y, data['y'])
#     x = np.append(x, data['x'])
    
#     lambda_index = np.argmax(y)
#     lambda_nm = x[lambda_index]
#     return lambda_index, lambda_nm #zwracam index z największą wartością nat. diody i długość fali

# lambdas_index = {}
# lambdas_nm = {}
# for key, value in figs.leds.items():
#     lambdas_index[key], lambdas_nm[key] = get_max_lambda(value)
#     #print(lambdas_nm[key])

# def get_trans_values(data):
#     y = np.array([])
#     y = np.append(y, data['y'])
#     vals = {}

#     for key, value in lambdas_index.items():
#         vals[key] = y[value] #dla każdego oleju pobieramy każdą wartość trans. dla indexu gdzie jest max nat. diody
#     return vals

# trans_vals = {}
# for key, value in figs.trans.items():
#     trans_vals[key] = get_trans_values(value)

# data = dict(
#     oils = ['Diesel (napędowy)', 'Castrol EDGE 0W-20 (silnikowy)', 
# 'Castrol EDGE 0W-30 (silnikowy)', 'Castrol MAGNATEC (silnikowy)', 'COMMA Pro-NRG (silnikowy)', 'COMMA Xtech (silnikowy)', 'elf EVOLUTION (silnikowy)', 'LOTOS (silnikowy)', 'MILLERS OILS (silnikowy)',
# 'Mobil Super 2000 (silnikowy)', 'MOTUL Specific (silnikowy)', 'COMMA ASW (przekładniowy)', 'febi bilstein axle drive (przekładniowy)', 'febi bilstein gear oil (przekładniowy)', 'jasol AFT (przekładniowy)',
# 'ORLEN HIPOL (przekładniowy)', 'jasol L-DAA 100 (sprężarkowy)', 'ORLEN BOXOL 26 (hydrauliczny)', 'TermoPasty (maszynowy)',
# 'Oliwa (spożywczy)', 'Rzepakowy (spożywczy)', 'Słonecznikowy (spożywczy)'],
#     blues = [trans_vals[key]['blue_spectrum'] for key, value in trans_vals.items()],
#     greens = [trans_vals[key]['green_spectrum'] for key, value in trans_vals.items()],
#     yellows = [trans_vals[key]['yellow_spectrum'] for key, value in trans_vals.items()],
#     oranges = [trans_vals[key]['orange_spectrum'] for key, value in trans_vals.items()],
#     reds = [trans_vals[key]['red_spectrum'] for key, value in trans_vals.items()],
#             )

# source = ColumnDataSource(data)

# def get_html_formatter(my_col):
#     template = """                
#                 <div style="background:<%=
#                                 (function colorfromint(){
#                                     if(col == blues)
#                                         {return('#7FB3D5')}
#                                     else if (col == greens)
#                                         {return('#73C6B6')}
#                                     else if (col == yellows)
#                                         {return('#F7DC6F')}
#                                     else if (col == oranges)
#                                         {return('#FFBC5D')}
#                                     else if (col == reds)
#                                         {return('#F1948A')}
#                                     }()) %>; 
#                                 color: black"> 
#                             <%= value %>    
#                 </div>
#                 """.replace('col',my_col)
#     return HTMLTemplateFormatter(template = template)

# columns = [
#         TableColumn(field = "oils", title = "Nazwa oleju"),
#         TableColumn(field = "blues", title = f"NIEBIESKA ({lambdas_nm['blue_spectrum']} nm) T[%]", formatter = get_html_formatter('blues')),
#         TableColumn(field = "greens", title = f"ZIELONA ({lambdas_nm['green_spectrum']} nm) T[%]", formatter = get_html_formatter('greens')),
#         TableColumn(field = "yellows", title = f"ŻÓŁTA ({lambdas_nm['yellow_spectrum']} nm) T[%]", formatter = get_html_formatter('yellows')),
#         TableColumn(field = "oranges", title = f"POMARAŃCZOWA ({lambdas_nm['orange_spectrum']} nm) T[%]", formatter = get_html_formatter('oranges')),
#         TableColumn(field = "reds", title = f"CZERWONA ({lambdas_nm['red_spectrum']} nm) T[%]", formatter = get_html_formatter('reds')),
#          ]

# trans_table = DataTable(source = source, columns = columns, width = 1200, height = 700)

############

def import_data(file_n, test_n, col1, col2, col3):
    wb = load_workbook(os.getcwd()+file_n)
    ws = wb.active

    if test_n == 'test1':
        start = '1'
        stop = '3'
    elif test_n == 'test2':
        start = '4'
        stop = '6'
    else:
        start = '7'
        stop = '9'

    range_1 = ws[col1+start:col1+stop]
    before = []
    range_2 = ws[col2+start:col2+stop]
    after = []
    range_3 = ws[col3+start:col3+stop]
    trans = []

    for cell in range_1:
        for e in cell:
            before.append(e.value)

    for cell in range_2:
        for e in cell:
            after.append(e.value)

    for cell in range_3:
        for e in cell:
            trans.append(e.value)

    leds = [1, 2, 3]

    data = {'leds': leds, 'before': before, 'after': after, 'trans': trans}
    return data

castrol_tests = defaultdict(dict)
jasol_tests = defaultdict(dict)
olive_tests = defaultdict(dict)

for mm in [['2', 'A', 'B', 'C'], ['4', 'E', 'F', 'G'], ['6', 'I', 'J', 'K'], ['8', 'M', 'N', 'O']]:
    for test in ['test1', 'test2', 'test3']:
        castrol_tests[mm[0]+'_mm'][f'{test}'] = import_data('\\results\\castrol.xlsx', test, mm[1], mm[2], mm[3])
        jasol_tests[mm[0]+'_mm'][f'{test}'] = import_data('\\results\\jasol.xlsx', test, mm[1], mm[2], mm[3])
        olive_tests[mm[0]+'_mm'][f'{test}'] = import_data('\\results\\olive.xlsx', test, mm[1], mm[2], mm[3])

def set_fig_properties(fig):
    #fig.background_fill_color = all_palettes['Greys'][9][7]
    fig.ygrid.grid_line_color = all_palettes['Greys'][9][6]
    fig.xgrid.grid_line_color = all_palettes['Greys'][9][6]
    fig.title.text_font_size = '16px'
    fig.title.text_font = 'Calibri Light'
    fig.yaxis.axis_label_text_font_size = '14px'
    fig.xaxis.axis_label_text_font_size = '14px'
    fig.yaxis.axis_label_text_font = 'Calibri Light'
    fig.xaxis.axis_label_text_font = 'Calibri Light'

#wykresy porównania powtarzalności pomiarów

fig_rep = defaultdict(dict)
line_rep = defaultdict(dict)

for fig in [['castrol', 'Castrol MAGNATEC', castrol_tests], ['jasol', 'jasol AFT', jasol_tests], ['olive', 'Olive', olive_tests]]:
    fig_rep[fig[0]] = figure(x_range = (0, 4.5), y_range = (0, 110), x_axis_label = 'Dioda', y_axis_label = 'Transmitancja [%]',
                    title = fig[1], width = 400, height = 300)
    set_fig_properties(fig_rep[fig[0]])
    fig_rep[fig[0]].xaxis.ticker = [1, 2, 3]
    fig_rep[fig[0]].xaxis.major_label_overrides = {1: 'niebieska', 2: 'zielona', 3: 'czerwona'}

    for mm in ['2', '4', '6', '8']:
        for test in [['test1', all_palettes['PuRd'][5][1]], ['test2', all_palettes['Bokeh'][6][5]], ['test3', all_palettes['Cividis'][3][1]]]:
            line_rep[test[0]] = fig_rep[fig[0]].circle('leds', 'trans', source = fig[2][mm+'_mm'][test[0]], size = 10, legend_label = f'{mm} mm', color = test[1])
            line_rep[test[0]].visible = False
            fig_rep[fig[0]].legend.click_policy = "hide"

#wykresy zależności trans od mm

def calculate_average(data, led):
    if led == 'niebieska':
        index = 0
    elif led == 'zielona':
        index = 1
    else:
        index = 2

    trans = np.array([])
    trans_avg = np.array([])

    for key, value in data.items():
        for test in ['test1', 'test2', 'test3']:
            data_new = data[key][test]
            data_trans = data_new['trans']
            trans = np.append(trans, data_trans[index])
        trans_avg = np.append(trans_avg, np.average(trans))
        trans = np.delete(trans, [0,1,2])
    
    t_avg = [trans_avg[0], trans_avg[1], trans_avg[2], trans_avg[3]]
    mm = [2, 4, 6, 8]
    return {'mm': mm, 'trans_avg': t_avg}

avg_castrol_trans = defaultdict(dict)
avg_jasol_trans = defaultdict(dict)
avg_olive_trans = defaultdict(dict)

for led in ['niebieska','zielona','czerwona']:
    avg_castrol_trans[led] = calculate_average(castrol_tests, led)
    avg_jasol_trans[led] = calculate_average(jasol_tests, led)
    avg_olive_trans[led] = calculate_average(olive_tests, led)

fig_trans = defaultdict(dict)
line_trans = defaultdict(dict)

for fig in [['castrol', 'Castrol MAGNATEC', avg_castrol_trans], ['jasol', 'jasol AFT', avg_jasol_trans], ['olive', 'Olive', avg_olive_trans]]:
    fig_trans[fig[0]] = figure(x_range = (0, 10), y_range = (0, 110), x_axis_label = 'Droga [mm]', y_axis_label = 'Transmitancja [%]',
                    title = fig[1], width = 400, height = 300)
    set_fig_properties(fig_trans[fig[0]])

    for led in [['niebieska', all_palettes['Spectral'][11][1]], ['zielona', all_palettes['PuBuGn'][5][0]], ['czerwona', all_palettes['Spectral'][11][10]]]:
        line_trans[led[0]] = fig_trans[fig[0]].circle('mm', 'trans_avg', source = fig[2][led[0]], size = 10, legend_label = led[0], color = led[1])

    fig_trans[fig[0]].legend.location = "bottom_left"
    fig_trans[fig[0]].legend.label_text_font_size = '8pt'
    fig_trans[fig[0]].legend.background_fill_alpha = 0.0
    #fig_trans[fig[0]].legend.visible = False

# legend_trans = Legend(items=
#                 [("niebieska", [line_trans['niebieska']]), 
#                 ("zielona", [line_trans['zielona']]), 
#                 ("czerwona", [line_trans['czerwona']])],
#                 location="top_left", orientation="horizontal")
# fig_trans['castrol'].add_layout(legend_trans, 'above')


#wykresy zależności stosunku trans diod od mm
def calculate_ratio(data1, data2):
    trans1 = data1['trans_avg']
    trans2 = data2['trans_avg']
    results = []

    for i in range(4):
        results.append(trans1[i]/trans2[i])

    return {'mm': data1['mm'], 'ratio': results}

ratio_castrol = defaultdict(dict)
ratio_jasol = defaultdict(dict)
ratio_olive = defaultdict(dict)

ratio_castrol['N/CZ'] = calculate_ratio(avg_castrol_trans['niebieska'], avg_castrol_trans['czerwona'])
ratio_castrol['Z/CZ'] = calculate_ratio(avg_castrol_trans['zielona'], avg_castrol_trans['czerwona'])
ratio_jasol['N/CZ'] = calculate_ratio(avg_jasol_trans['niebieska'], avg_jasol_trans['czerwona'])
ratio_jasol['Z/CZ'] = calculate_ratio(avg_jasol_trans['zielona'], avg_jasol_trans['czerwona'])
ratio_olive['N/Z'] = calculate_ratio(avg_olive_trans['niebieska'], avg_olive_trans['zielona'])
ratio_olive['CZ/Z'] = calculate_ratio(avg_olive_trans['czerwona'], avg_olive_trans['zielona'])

fig_ratio = defaultdict(dict)
circle_ratio = defaultdict(dict)

for fig in [['castrol', 'Castrol MAGNATEC', ratio_castrol], ['jasol', 'jasol AFT', ratio_jasol], ['olive', 'Olive', ratio_olive]]:
    fig_ratio[fig[0]] = figure(x_range = (0, 10), y_range = (0, 1.1), x_axis_label = 'Droga [mm]', y_axis_label = 'Stosunek',
                    title = fig[1], width = 400, height = 300)
    set_fig_properties(fig_ratio[fig[0]])

    # for ratio1 in [['N/CZ', all_palettes['PuRd'][5][1]], ['Z/CZ', all_palettes['Bokeh'][6][5]]]:
    #     line_ratio[ratio1[0]] = fig_ratio[fig[0]].circle('mm', 'ratio', source = fig[2][ratio1[0]], size = 10, legend_label = ratio1[0], color = ratio1[1])
    
def set_legend(fig):
    fig.legend.location = "bottom_left"
    fig.legend.label_text_font_size = '8pt'
    fig.legend.background_fill_alpha = 0.0

circle_ratio['N/CZ_castrol'] = fig_ratio['castrol'].circle('mm', 'ratio', source = ratio_castrol['N/CZ'], size = 10, legend_label = 'N/CZ', color = all_palettes['PuRd'][5][1])
circle_ratio['Z/CZ_castrol'] = fig_ratio['castrol'].circle('mm', 'ratio', source = ratio_castrol['Z/CZ'], size = 10, legend_label = 'Z/CZ', color = all_palettes['Bokeh'][6][5])
circle_ratio['N/CZ_jasol'] = fig_ratio['jasol'].circle('mm', 'ratio', source = ratio_jasol['N/CZ'], size = 10, legend_label = 'N/CZ', color = all_palettes['PuRd'][5][1])
circle_ratio['Z/CZ_jasol'] = fig_ratio['jasol'].circle('mm', 'ratio', source = ratio_jasol['Z/CZ'], size = 10, legend_label = 'Z/CZ', color = all_palettes['Bokeh'][6][5])
circle_ratio['N/Z_olive'] = fig_ratio['olive'].circle('mm', 'ratio', source = ratio_olive['N/Z'], size = 10, legend_label = 'N/Z', color = all_palettes['PuRd'][5][1])
circle_ratio['CZ/Z_olive'] = fig_ratio['olive'].circle('mm', 'ratio', source = ratio_olive['CZ/Z'], size = 10, legend_label = 'CZ/Z', color = all_palettes['Bokeh'][6][5])

set_legend(fig_ratio['castrol'])
set_legend(fig_ratio['jasol'])
set_legend(fig_ratio['olive'])

#dopasowanie krzywej
def linear_reg(data):
    x = np.array([])
    y = np.array([])
    x = np.append(x, data['mm'])
    y = np.append(y, data['ratio'])

    linreg = LinearRegression()
    x = x.reshape(-1,1)
    linreg.fit(x, y)
    y_pred = linreg.predict(x)

    return {'x': x, 'y_pred': y_pred}, linreg.coef_, linreg.intercept_

linear_castrol = defaultdict(dict)
linear_jasol = defaultdict(dict)
linear_olive = defaultdict(dict)
a_castrol = defaultdict(dict)
b_castrol = defaultdict(dict)
a_jasol = defaultdict(dict)
b_jasol = defaultdict(dict)
a_olive = defaultdict(dict)
b_olive = defaultdict(dict)

linear_castrol['N/CZ'], a_castrol['N/CZ'], b_castrol['N/CZ'] = linear_reg(ratio_castrol['N/CZ'])
linear_castrol['Z/CZ'], a_castrol['Z/CZ'], b_castrol['Z/CZ'] = linear_reg(ratio_castrol['Z/CZ'])
linear_jasol['N/CZ'], a_jasol['N/CZ'], b_jasol['N/CZ'] = linear_reg(ratio_jasol['N/CZ'])
linear_jasol['Z/CZ'], a_jasol['Z/CZ'], b_jasol['Z/CZ'] = linear_reg(ratio_jasol['Z/CZ'])
linear_olive['N/Z'], a_olive['N/Z'], b_olive['N/Z'] = linear_reg(ratio_olive['N/Z'])
linear_olive['CZ/Z'], a_olive['CZ/Z'], b_olive['CZ/Z'] = linear_reg(ratio_olive['CZ/Z'])

line_ratio = defaultdict(dict)

line_ratio['N/CZ_castrol'] = fig_ratio['castrol'].line('x', 'y_pred', source = linear_castrol['N/CZ'], legend_label = 'N/CZ', color = all_palettes['PuRd'][5][1])
line_ratio['Z/CZ_castrol'] = fig_ratio['castrol'].line('x', 'y_pred', source = linear_castrol['Z/CZ'], legend_label = 'Z/CZ', color = all_palettes['Bokeh'][6][5])
line_ratio['N/CZ_jasol'] = fig_ratio['jasol'].line('x', 'y_pred', source = linear_jasol['N/CZ'], legend_label = 'N/CZ', color = all_palettes['PuRd'][5][1])
line_ratio['Z/CZ_jasol'] = fig_ratio['jasol'].line('x', 'y_pred', source = linear_jasol['Z/CZ'], legend_label = 'Z/CZ', color = all_palettes['Bokeh'][6][5])
line_ratio['N/Z_olive'] = fig_ratio['olive'].line('x', 'y_pred', source = linear_olive['N/Z'], legend_label = 'N/Z', color = all_palettes['PuRd'][5][1])
line_ratio['CZ/Z_olive'] = fig_ratio['olive'].line('x', 'y_pred', source = linear_olive['CZ/Z'], legend_label = 'CZ/Z', color = all_palettes['Bokeh'][6][5])

div_reps = Div(text = 'Wykresy przedstawiające powtarzalność pomiarów:')
div_trans = Div(text = 'Wykresy przedstawiające zależność transmitancji od drogi:')
div_ratio = Div(text = 'Wykresy przedstawiające zależność stosunku transmitancji między poszczególnymi diodami od drogi:')

with open(os.getcwd()+'\\results.txt', 'w') as file:
    for data in [['castrol N/CZ', a_castrol['N/CZ'], b_castrol['N/CZ']], ['castrol Z/CZ', a_castrol['Z/CZ'], b_castrol['Z/CZ']], ['jasol N/CZ', a_jasol['N/CZ'], b_jasol['N/CZ']],
                ['jasol Z/CZ', a_jasol['Z/CZ'], b_jasol['Z/CZ']], ['olive N/Z', a_olive['N/Z'], b_olive['N/Z']], ['olive CZ/Z', a_olive['CZ/Z'], b_olive['CZ/Z']]]:
        file.write(f'{data[0]}, {data[1]}, {data[2]}\n')